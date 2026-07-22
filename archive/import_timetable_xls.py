import argparse
import json
import re
from datetime import datetime
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

DAYS = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
BREAK_MARKERS = {"R", "E", "C", "S", "BREAK", "RECESS"}
FREE_MARKERS = {"", "FREE", "VACANT SLOT", "VACANT", "NIL", "---", "NA"}


def clean_text(value):
    text = str(value or "")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() == "nan":
        return ""
    return text


def slugify(value):
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return slug or "room"


def normalize_room_code(value):
    room = clean_text(value).upper().replace(" ", "")
    return room.strip("-")


def parse_wef_date(header_line):
    match = re.search(r"W\.?\s*E\.?\s*F\.?\s*[:-]\s*([0-9]{1,2}\s*/\s*[0-9]{1,2}\s*/\s*[0-9]{2,4})", header_line, flags=re.IGNORECASE)
    if not match:
        return None

    raw_date = re.sub(r"\s+", "", match.group(1))
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw_date, fmt)
        except ValueError:
            continue
    return None


class XlrdSheetAdapter:
    def __init__(self, sheet):
        self.sheet = sheet
        self.name = sheet.name
        self.nrows = sheet.nrows
        self.ncols = sheet.ncols
        self.merged_ranges = sheet.merged_cells

    def get_direct_text(self, row_index, col_index):
        if row_index < 0 or row_index >= self.nrows or col_index < 0 or col_index >= self.ncols:
            return ""
        return clean_text(self.sheet.cell_value(row_index, col_index))

    def get_cell_text(self, row_index, col_index):
        direct = self.get_direct_text(row_index, col_index)
        if direct:
            return direct

        for row_start, row_end, col_start, col_end in self.merged_ranges:
            if row_start <= row_index < row_end and col_start <= col_index < col_end:
                return self.get_direct_text(row_start, col_start)

        return ""


class OpenpyxlSheetAdapter:
    def __init__(self, sheet):
        self.sheet = sheet
        self.name = sheet.title
        self.nrows = sheet.max_row
        self.ncols = sheet.max_column
        self.merged_ranges = []

        for cell_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
            self.merged_ranges.append((min_row - 1, max_row, min_col - 1, max_col))

    def get_direct_text(self, row_index, col_index):
        if row_index < 0 or row_index >= self.nrows or col_index < 0 or col_index >= self.ncols:
            return ""
        return clean_text(self.sheet.cell(row=row_index + 1, column=col_index + 1).value)

    def get_cell_text(self, row_index, col_index):
        direct = self.get_direct_text(row_index, col_index)
        if direct:
            return direct

        for row_start, row_end, col_start, col_end in self.merged_ranges:
            if row_start <= row_index < row_end and col_start <= col_index < col_end:
                return self.get_direct_text(row_start, col_start)

        return ""


def load_workbook_sheets(input_path):
    suffix = input_path.suffix.lower()

    if suffix == ".xls":
        workbook = xlrd.open_workbook(str(input_path), formatting_info=False)
        return [XlrdSheetAdapter(sheet) for sheet in workbook.sheets()]

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(filename=str(input_path), data_only=True)
        return [OpenpyxlSheetAdapter(sheet) for sheet in workbook.worksheets]

    raise ValueError(f"Unsupported workbook format: {input_path}")


def parse_header_metadata(header_line, sheet_name):
    class_match = re.search(
        r"CLASS\s*:\s*(.*?)(?:Lect\.?\s*Hall\s*No\.?-|Class\s*room:|Room\s*No\.?|TIME\s+TABLE|W\.?\s*E\.?\s*F\.?|$)",
        header_line,
        flags=re.IGNORECASE,
    )

    if class_match:
        class_name = clean_text(class_match.group(1))
    else:
        lab_match = re.search(
            r"Name\s*of\s*Lab\s*:\s*(.*?)(?:\(\s*Room\s*No\.?|TIME\s+TABLE|W\.?\s*E\.?\s*F\.?|$)",
            header_line,
            flags=re.IGNORECASE,
        )
        class_name = clean_text(lab_match.group(1)) if lab_match else ""

    room_match = re.search(
        r"(?:Lect\.?\s*Hall\s*No\.?-|Class\s*room:|Room\s*No\.?\s*[:\.-]?)\s*([A-Za-z0-9\-]+)",
        header_line,
        flags=re.IGNORECASE,
    )
    room_code = normalize_room_code(room_match.group(1)) if room_match else ""

    if not class_name:
        class_name = clean_text(sheet_name.replace("_", " "))

    if not room_code and re.match(r"^\d+[A-Z]?$", normalize_room_code(sheet_name)):
        room_code = normalize_room_code(sheet_name)

    return class_name, room_code


def normalize_time(raw_time):
    value = clean_text(raw_time).lower().replace(".", ":")
    match = re.search(r"(\d{1,2})(?::(\d{1,2}))?\s*([ap]m)?", value)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    period = match.group(3)

    if period == "pm" and hour < 12:
        hour += 12
    elif period == "am" and hour == 12:
        hour = 0
    elif not period and 1 <= hour <= 6:
        hour += 12

    return f"{hour:02d}:{minute:02d}"


def normalize_slot(raw_slot):
    value = clean_text(raw_slot)
    if not value or value.upper() == "TIMING":
        return None

    parts = re.findall(r"\d{1,2}(?:[\.:]\d{1,2})?(?:\s*[ap]m)?", value, flags=re.IGNORECASE)
    if len(parts) < 2:
        return None

    start = normalize_time(parts[0])
    end = normalize_time(parts[1])
    if not start or not end:
        return None

    return f"{start}-{end}"


def normalize_subject(subject_raw):
    subject = clean_text(subject_raw)
    if not subject:
        return "Free"

    upper = subject.upper()
    alpha_compact = re.sub(r"[^A-Z]", "", upper)

    if upper in BREAK_MARKERS or alpha_compact == "RECESS":
        return "Break"

    if upper in FREE_MARKERS or upper.startswith("VACANT SLOT"):
        return "Free"

    return subject


def find_header_line(sheet):
    markers = ("CLASS:", "LECT. HALL NO", "CLASS ROOM", "NAME OF LAB", "ROOM NO")
    candidates = []

    max_rows = min(sheet.nrows, 14)
    max_cols = min(sheet.ncols, 4)

    for row_index in range(max_rows):
        for col_index in range(max_cols):
            value = sheet.get_cell_text(row_index, col_index)
            if any(marker in value.upper() for marker in markers):
                candidates.append(value)

    if candidates:
        return max(candidates, key=len)

    return sheet.get_cell_text(3, 1) or sheet.get_cell_text(3, 0)


def find_timing_row(sheet):
    max_rows = min(sheet.nrows, 40)
    max_cols = min(sheet.ncols, 4)

    for row_index in range(max_rows):
        for col_index in range(max_cols):
            value = sheet.get_cell_text(row_index, col_index).upper()
            if value == "TIMING":
                return row_index

    return None


def find_day_rows(sheet):
    day_rows = {}
    max_cols = min(sheet.ncols, 3)

    for row_index in range(sheet.nrows):
        for col_index in range(max_cols):
            day = sheet.get_cell_text(row_index, col_index).upper()
            if day in DAYS and day not in day_rows:
                day_rows[day] = row_index

    return day_rows


def find_slot_columns(sheet, timing_row_index):
    slot_columns = []

    for col_index in range(sheet.ncols):
        slot = normalize_slot(sheet.get_direct_text(timing_row_index, col_index))
        if not slot:
            continue

        if slot_columns and slot_columns[-1][1] == slot:
            continue

        slot_columns.append((col_index, slot))

    return slot_columns


def extract_sheet(sheet, source_file_name):
    header_line = find_header_line(sheet)
    class_name, room_code = parse_header_metadata(header_line, sheet.name)

    if not room_code:
        return None

    timing_row_index = find_timing_row(sheet)
    if timing_row_index is None:
        return None

    slot_columns = find_slot_columns(sheet, timing_row_index)
    if not slot_columns:
        return None

    day_rows = find_day_rows(sheet)
    if len(day_rows) < 5:
        return None

    weekly_timetable = {}
    meaningful_count = 0

    for day in DAYS:
        row_index = day_rows.get(day)
        if row_index is None:
            continue

        day_slots = []
        for col_index, slot in slot_columns:
            subject = normalize_subject(sheet.get_cell_text(row_index, col_index))

            if subject not in {"Free", "Break"}:
                meaningful_count += 1

            day_slots.append(
                {
                    "slot": slot,
                    "subject": subject,
                    "faculty": "",
                }
            )

        weekly_timetable[day] = day_slots

    if meaningful_count == 0:
        return None

    room_id = slugify(room_code)
    room_type = "lab" if "LAB" in header_line.upper() or "LAB" in class_name.upper() else "classroom"

    room = {
        "id": room_id,
        "name": room_code,
        "type": room_type,
        "capacity": 0,
        "equipment": [],
        "weeklyTimetable": weekly_timetable,
        "timetable": weekly_timetable.get("MONDAY", []),
        "sourceClass": class_name,
        "sourceFile": source_file_name,
        "location": room_code,
    }

    duplicate_penalty = 0 if "(2)" in sheet.name else 1
    wef_date = parse_wef_date(header_line)
    wef_rank = int(wef_date.timestamp()) if wef_date else 0

    return {
        "room_key": room_code,
        "sheet_name": sheet.name,
        "score": (meaningful_count, wef_rank, duplicate_penalty),
        "room": room,
    }


def room_sort_key(room):
    match = re.match(r"^(\d+)([A-Z]*)$", room["name"].upper())
    if match:
        return (int(match.group(1)), match.group(2))
    return (9999, room["name"].upper())


def build_room_page_html(template_html, room):
    html = re.sub(
        r"<title>.*?</title>",
        f"<title>Room {room['name']} | Smart Campus</title>",
        template_html,
        count=1,
        flags=re.IGNORECASE | re.DOTALL,
    )

    def replace_body(match):
        attrs = match.group(1) or ""
        attrs = re.sub(r'\sdata-room-id="[^"]*"', "", attrs, flags=re.IGNORECASE)
        return f'<body{attrs} data-room-id="{room["id"]}">'

    return re.sub(r"<body([^>]*)>", replace_body, html, count=1, flags=re.IGNORECASE)


def write_room_pages(rooms, template_path, output_dir):
    template_html = template_path.read_text(encoding="utf-8")

    for old_page in output_dir.glob("room-*.html"):
        old_page.unlink()

    generated = []
    for room in rooms:
        page_filename = f"room-{room['id']}.html"
        room["page"] = page_filename

        page_html = build_room_page_html(template_html, room)
        page_path = output_dir / page_filename
        page_path.write_text(page_html, encoding="utf-8")
        generated.append(page_filename)

    return generated


def convert_workbooks(input_paths, output_path, room_template_path, pages_output_dir):
    selected_by_room = {}

    for input_index, input_path in enumerate(input_paths):
        for sheet in load_workbook_sheets(input_path):
            parsed = extract_sheet(sheet, input_path.name)
            if not parsed:
                continue

            room_key = parsed["room_key"]
            parsed["score"] = parsed["score"] + (input_index,)

            current = selected_by_room.get(room_key)
            if not current or parsed["score"] > current["score"]:
                selected_by_room[room_key] = parsed

    rooms = [item["room"] for item in selected_by_room.values()]
    rooms.sort(key=room_sort_key)

    # Disable static page generation
    # generated_pages = write_room_pages(rooms, room_template_path, pages_output_dir)

    output_path.write_text(json.dumps(rooms, indent=2), encoding="utf-8")

    input_names = ", ".join(path.name for path in input_paths)
    print(f"Imported {len(rooms)} schedules from: {input_names}")
    print(f"Wrote {output_path.name} and generated {len(generated_pages)} room pages in {pages_output_dir}")

    for item in sorted(selected_by_room.values(), key=lambda value: room_sort_key(value["room"])):
        room = item["room"]
        print(
            f"- {item['sheet_name']} ({room['sourceFile']}) -> Room {room['name']} "
            f"(from {room['sourceClass']}), page room-{room['id']}.html"
        )


def default_input_files():
    candidates = []
    for pattern in ("*.xls", "*.xlsx", "*.xlsm", "*.xltx", "*.xltm"):
        candidates.extend(sorted(Path(".").glob(pattern)))
    return [str(path) for path in candidates]


def main():
    parser = argparse.ArgumentParser(
        description="Convert one or more timetable workbooks into room-wise data.json and dedicated room pages"
    )
    parser.add_argument(
        "--input",
        nargs="+",
        default=default_input_files(),
        help="Input timetable workbook file(s): .xls or .xlsx",
    )
    parser.add_argument(
        "--output",
        default="data.json",
        help="Output JSON file path",
    )
    parser.add_argument(
        "--room-template",
        default="room.html",
        help="Template HTML file used to generate one file per room",
    )
    parser.add_argument(
        "--pages-dir",
        default=".",
        help="Directory where generated room-*.html files are written",
    )
    args = parser.parse_args()

    if not args.input:
        raise FileNotFoundError("No input workbook files found. Add .xls/.xlsx files or pass --input explicitly.")

    input_paths = [Path(path) for path in args.input]
    output_path = Path(args.output)
    room_template_path = Path(args.room_template)
    pages_output_dir = Path(args.pages_dir)

    missing_inputs = [str(path) for path in input_paths if not path.exists()]
    if missing_inputs:
        raise FileNotFoundError(f"Input file(s) not found: {', '.join(missing_inputs)}")

    if not room_template_path.exists():
        raise FileNotFoundError(f"Room template file not found: {room_template_path}")

    pages_output_dir.mkdir(parents=True, exist_ok=True)

    convert_workbooks(input_paths, output_path, room_template_path, pages_output_dir)


if __name__ == "__main__":
    main()
